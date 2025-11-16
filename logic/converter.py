import openpyxl
import csv
import re
from openpyxl.utils import get_column_letter


# Concerned columns are:
COL1 = "stock central depart"
COL2 = "colline"


# Loader class
# @brief: This class loads all the stuff this program needs to run
class Loader:
    def __init__(self, coordinates_path: str) -> None:
        self.coord_path = coordinates_path  # Must be in csv format
        self.file = open(self.coord_path, newline='')
        self.csv_object = csv.reader(self.file, delimiter=',', quotechar='"')

    def test(self):
        for row in self.csv_object:
            print(row)

    def close(self):
        self.file.close()


# Reader class
# @brief: This class reads the file given by the converter and processes it
class Reader:
    def __init__(self, excel_path: str) -> None:
        self.file_path = excel_path
        self.workbook = openpyxl.load_workbook(self.file_path)
        self.sheet = self.workbook.active
        self.rows = self.sheet.max_row
        self.columns = self.sheet.max_column

        next_col = self.columns + 1
        col_letter = get_column_letter(next_col)
        col_letter_next = get_column_letter(next_col+1)

        self.sheet[f"{col_letter}1"] = "From"
        self.sheet[f"{col_letter_next}1"] = "To"

        self.workbook.save("/home/jaspe/hello.xlsx")

        self.columns += 2
        self.map: dict = dict()  # Map containing a certain field(string) with it's column number

    def complete_map(self):
        for i in range(1, self.columns+1):
            cell = self.sheet.cell(row=1, column=i)
            if cell.value is not None:
                self.map[cell.value.lower().strip()] = i

    def get_column(self, name: str) -> int:
        return self.map.get(name, 0)

    def get_row(self, row: int) -> list:
        assert 1 <= row <= self.rows

        cells = []

        for i in range(1, self.columns+1):
            cell = self.sheet.cell(row=row, column=i)
            cells.append(cell.value)

        return cells

    def write_from_to(self, row: int, _from: str, _to: str) -> None:
        from_col, to_col = self.map["from"], self.map["to"]
        from_letter, to_letter = get_column_letter(from_col), \
                get_column_letter(to_col)
        self.sheet[f"{from_letter}{row}"] = _from
        self.sheet[f"{to_letter}{row}"] = _to


class Converter:
    def __init__(self, excel_path: str, coordinates_path: str) -> None:
        self.loader: Loader = Loader(coordinates_path)
        self.reader: Reader = Reader(excel_path)
        self.reader.complete_map()

    def purify_stocks(self, _stock: str) -> str:
        stock = _stock.lower().strip()
        fmt = r"stock.+central(.+)"
        new_name = re.split(fmt, stock)[1].strip()
        return new_name.capitalize()

    def test(self) -> None:
        print(self.reader.get_row(1))
        print(self.reader.get_row(2))

    # Link the tuple (a,b) to a string 'a_b'
    def link_from_to(self, from_to: tuple) -> str:
        return '_'.join(from_to)

    def get_from_to(self) -> dict:
        # print(self.reader.get_row(8))
        # for row in range(1, self.reader.rows+1):
        depart, colline = self.reader.get_column("stock central depart")-1, \
                self.reader.get_column("colline")-1
        numero_mvt = self.reader.get_column("numero du mouvement")-1
        district = self.reader.get_column("district")-1

        from_to: dict = dict()

        current_row = 2
        reading_row = 1
        last_colline = ""
        last_numero_mvt = -1
        for _row in range(current_row, 16):
            row = self.reader.get_row(_row)
            if row[numero_mvt] != last_numero_mvt:
                reading_row = 1
                last_colline = ""
                last_numero_mvt = row[numero_mvt]

            # If row is empty, skip it
            if row[numero_mvt] is None:
                continue

            _colline = row[colline]
            if reading_row == 1:
                _stock = self.purify_stocks(row[depart])
                from_to[last_numero_mvt] = from_to.get(last_numero_mvt, []) +\
                        [(self.purify_stocks(row[depart]), colline)]
                self.reader.write_from_to(_row, _stock, _colline)
            else:
                from_to[last_numero_mvt] = from_to.get(last_numero_mvt, []) + [(last_colline, colline)]
                self.reader.write_from_to(_row, last_colline, _colline)

            last_colline = row[colline]
            current_row += 1
            reading_row += 1
        self.reader.workbook.save("/home/jaspe/hello.xlsx")

        return from_to


if __name__ == "__main__":
    exit(0)
